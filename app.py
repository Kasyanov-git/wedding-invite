# -*- coding: utf-8 -*-
"""
Wedding Invitation Backend
==========================
Flask-приложение для сбора ответов гостей и выгрузки в Excel.

Переменные окружения:
    ADMIN_PASSWORD    — пароль для доступа к /admin (по умолчанию: wedding2025)
    DATABASE_PATH     — путь к SQLite (по умолчанию: wedding.db)
    HOST / PORT       — настройки запуска
"""

import os
import sqlite3
import csv
import io
from datetime import datetime
from functools import wraps

from flask import (
    Flask, request, jsonify, send_file, Response, render_template_string,
    session, redirect, url_for, flash
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ───────────────────────────────────────────
# Конфигурация
# ───────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DATABASE_PATH", os.path.join(BASE_DIR, "wedding.db"))
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "wedding2025")

app = Flask(__name__, static_folder="static", static_url_path="")
app.secret_key = os.environ.get("SECRET_KEY", "wedding-invite-secret-key-change-me")

# ───────────────────────────────────────────
# Шаблон страницы входа в админку
# ───────────────────────────────────────────
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Вход в админку — Свадебное приглашение</title>
    <style>
        :root {
            --cream: #FBF7F1;
            --powder: #F3E5E0;
            --nude: #E6C8B3;
            --beige: #D4B896;
            --caramel: #B08D6F;
            --coffee: #6B4E3D;
            --text: #4A3F3A;
        }
        * { box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            background: linear-gradient(135deg, var(--cream) 0%, var(--powder) 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            margin: 0;
            padding: 1rem;
            color: var(--text);
        }
        .login-card {
            background: rgba(255, 255, 255, 0.75);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border-radius: 32px;
            border: 1px solid rgba(255, 255, 255, 0.6);
            box-shadow: 0 20px 60px rgba(107, 78, 61, 0.12);
            padding: 2.5rem 2rem;
            width: 100%;
            max-width: 420px;
            text-align: center;
        }
        .login-icon {
            width: 72px;
            height: 72px;
            border-radius: 50%;
            background: linear-gradient(145deg, var(--caramel), var(--coffee));
            color: white;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 2rem;
            margin: 0 auto 1.5rem;
            box-shadow: 0 8px 24px rgba(107, 78, 61, 0.2);
        }
        h1 {
            font-family: Georgia, serif;
            font-size: 1.5rem;
            color: var(--coffee);
            margin: 0 0 0.5rem;
            font-weight: 500;
        }
        p.subtitle {
            color: var(--caramel);
            font-size: 0.95rem;
            margin-bottom: 1.8rem;
        }
        .form-group {
            margin-bottom: 1.2rem;
            text-align: left;
        }
        label {
            display: block;
            font-size: 0.9rem;
            color: var(--coffee);
            margin-bottom: 0.4rem;
            font-weight: 500;
        }
        input[type="password"] {
            width: 100%;
            padding: 0.9rem 1rem;
            border: 1px solid var(--nude);
            border-radius: 16px;
            background: rgba(255, 255, 255, 0.8);
            font-size: 1rem;
            color: var(--text);
            outline: none;
            transition: border-color 0.2s, box-shadow 0.2s;
        }
        input[type="password"]:focus {
            border-color: var(--caramel);
            box-shadow: 0 0 0 3px rgba(176, 141, 111, 0.15);
        }
        button {
            width: 100%;
            padding: 1rem;
            border: none;
            border-radius: 50px;
            background: linear-gradient(145deg, var(--caramel) 0%, var(--coffee) 100%);
            color: white;
            font-size: 1rem;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s, box-shadow 0.2s;
            box-shadow: 0 4px 16px rgba(107, 78, 61, 0.2);
            margin-top: 0.5rem;
        }
        button:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(107, 78, 61, 0.25); }
        .error {
            background: rgba(237, 222, 222, 0.8);
            color: #7A3A3A;
            padding: 0.8rem 1rem;
            border-radius: 12px;
            font-size: 0.9rem;
            margin-bottom: 1rem;
        }
    </style>
</head>
<body>
    <div class="login-card">
        <div class="login-icon">❦</div>
        <h1>Админка приглашения</h1>
        <p class="subtitle">Вход для молодожёнов</p>
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        <form method="POST" action="/login">
            <div class="form-group">
                <label for="password">Пароль</label>
                <input type="password" id="password" name="password" required autofocus placeholder="Введите пароль">
            </div>
            <button type="submit">Войти</button>
        </form>
    </div>
</body>
</html>
"""

# ───────────────────────────────────────────
# Инициализация базы данных
# ───────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS rsvp (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            attendance TEXT NOT NULL,
            guests_count INTEGER DEFAULT 1,
            guest_names TEXT,
            second_day TEXT,
            drinks TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    # Миграции: добавляем колонки, если их нет (для уже созданных БД)
    for column in ["guest_names", "second_day"]:
        try:
            cursor.execute(f"ALTER TABLE rsvp ADD COLUMN {column} TEXT")
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()


init_db()

# ───────────────────────────────────────────
# Вспомогательные функции
# ───────────────────────────────────────────
def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def require_auth(f):
    """Проверка сессии для админки."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ───────────────────────────────────────────
# Публичные маршруты
# ───────────────────────────────────────────
@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("admin_logged_in"):
        return redirect(url_for("admin_panel"))

    error = None
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        if password == ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_panel"))
        error = "Неверный пароль. Попробуйте ещё раз."

    return render_template_string(LOGIN_TEMPLATE, error=error)


@app.route("/logout")
def logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("login"))


# ───────────────────────────────────────────
# API маршруты
# ───────────────────────────────────────────
@app.route("/api/rsvp", methods=["POST"])
def submit_rsvp():
    data = request.get_json(force=True)

    # Валидация обязательных полей
    full_name = (data.get("full_name") or "").strip()
    attendance = (data.get("attendance") or "").strip()

    if not full_name or not attendance:
        return jsonify({"success": False, "error": "Укажите ФИО и подтверждение участия"}), 400

    try:
        guests_count = max(1, int(data.get("guests_count") or 1))
    except (ValueError, TypeError):
        guests_count = 1

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO rsvp (full_name, attendance, guests_count, guest_names, second_day, drinks, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            full_name,
            attendance,
            guests_count,
            (data.get("guest_names") or "").strip(),
            (data.get("second_day") or "").strip(),
            (data.get("drinks") or "").strip(),
            now_iso(),
        ),
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "Ответ сохранён, спасибо!"})


@app.route("/api/rsvp", methods=["GET"])
@require_auth
def list_rsvp():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM rsvp ORDER BY created_at DESC")
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/admin/export")
@require_auth
def export_excel():
    """Выгрузка всех ответов в Excel (.xlsx)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM rsvp ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ответы гостей"

    headers = [
        "ID",
        "ФИО",
        "Участие",
        "Количество гостей",
        "Имена гостей",
        "Второй день",
        "Напитки",
        "Дата отправки",
    ]

    # Стили заголовков
    header_fill = PatternFill(start_color="E6C8B3", end_color="E6C8B3", fill_type="solid")
    header_font = Font(bold=True, color="4A3F3A")
    thin_border = Border(
        left=Side(style="thin", color="B08D6F"),
        right=Side(style="thin", color="B08D6F"),
        top=Side(style="thin", color="B08D6F"),
        bottom=Side(style="thin", color="B08D6F"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        values = [
            row["id"],
            row["full_name"],
            row["attendance"],
            row["guests_count"],
            row["guest_names"],
            row["second_day"],
            row["drinks"],
            row["created_at"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # Заморозка заголовка
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"wedding_rsvp_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/admin/export/csv")
@require_auth
def export_csv():
    """Выгрузка всех ответов в CSV."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM rsvp ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "ФИО", "Участие", "Количество гостей", "Имена гостей", "Второй день", "Напитки", "Дата отправки"
    ])
    for row in rows:
        writer.writerow([
            row["id"], row["full_name"], row["attendance"], row["guests_count"],
            row["guest_names"], row["second_day"], row["drinks"], row["created_at"]
        ])

    response = Response(output.getvalue(), mimetype="text/csv; charset=utf-8-sig")
    filename = f"wedding_rsvp_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.csv"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/admin")
@require_auth
def admin_panel():
    """Простая HTML-админка для просмотра и выгрузки ответов."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM rsvp ORDER BY created_at DESC")
    rows = cursor.fetchall()
    conn.close()

    total = len(rows)
    confirmed = sum(1 for r in rows if r["attendance"] == "confirmed")
    declined = sum(1 for r in rows if r["attendance"] == "declined")
    maybe = sum(1 for r in rows if r["attendance"] == "maybe")
    total_guests = sum(r["guests_count"] for r in rows if r["attendance"] == "confirmed")

    html = """
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Админка — ответы гостей</title>
        <style>
            :root {
                --cream: #FBF7F1;
                --nude: #E6C8B3;
                --beige: #D4B896;
                --caramel: #B08D6F;
                --coffee: #6B4E3D;
                --text: #4A3F3A;
            }
            * { box-sizing: border-box; }
            body {
                font-family: 'Segoe UI', system-ui, sans-serif;
                background: var(--cream);
                color: var(--text);
                margin: 0;
                padding: 2rem 1rem;
                line-height: 1.5;
            }
            h1 { margin-top: 0; color: var(--coffee); }
            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
                gap: 1rem;
                margin-bottom: 2rem;
            }
            .stat-card {
                background: white;
                border-radius: 20px;
                padding: 1.2rem;
                text-align: center;
                border: 1px solid var(--nude);
            }
            .stat-card .number {
                font-size: 2rem;
                font-weight: 700;
                color: var(--caramel);
            }
            .stat-card .label { font-size: 0.9rem; color: var(--coffee); }
            .actions { margin-bottom: 1.5rem; }
            .btn {
                display: inline-block;
                padding: 0.8rem 1.4rem;
                border-radius: 50px;
                background: var(--caramel);
                color: white;
                text-decoration: none;
                font-weight: 600;
                margin-right: 0.5rem;
                margin-bottom: 0.5rem;
                transition: transform .2s, background .2s;
            }
            .btn:hover { background: var(--coffee); transform: translateY(-2px); }
            table {
                width: 100%;
                border-collapse: collapse;
                background: white;
                border-radius: 20px;
                overflow: hidden;
                box-shadow: 0 4px 20px rgba(74,63,58,0.06);
            }
            th, td {
                padding: 0.8rem;
                text-align: left;
                border-bottom: 1px solid var(--nude);
                vertical-align: top;
            }
            th {
                background: var(--nude);
                color: var(--coffee);
                font-weight: 600;
            }
            tr:last-child td { border-bottom: none; }
            .badge {
                display: inline-block;
                padding: 0.25rem 0.7rem;
                border-radius: 50px;
                font-size: 0.85rem;
                font-weight: 600;
            }
            .badge-confirmed { background: #D4E9D7; color: #2D5A3D; }
            .badge-declined { background: #EDDEDE; color: #7A3A3A; }
            .badge-maybe { background: #F3E9D2; color: #8C6A3D; }
            .empty { text-align: center; padding: 3rem; color: var(--coffee); }
            @media (max-width: 768px) {
                table { display: block; overflow-x: auto; }
                body { padding: 1rem 0.5rem; }
            }
        </style>
    </head>
    <body>
        <h1>📋 Ответы гостей на свадьбу</h1>

        <div class="stats">
            <div class="stat-card">
                <div class="number">{{ total }}</div>
                <div class="label">Всего ответов</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ confirmed }}</div>
                <div class="label">Придут</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ declined }}</div>
                <div class="label">Не придут</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ maybe }}</div>
                <div class="label">Возможно</div>
            </div>
            <div class="stat-card">
                <div class="number">{{ total_guests }}</div>
                <div class="label">Ожидается гостей</div>
            </div>
        </div>

        <div class="actions">
            <a class="btn" href="/admin/export">⬇ Скачать Excel</a>
            <a class="btn" href="/logout" style="background: transparent; color: var(--caramel); border: 1px solid var(--caramel);">🚪 Выйти</a>
        </div>

        {% if rows %}
        <table>
            <thead>
                <tr>
                    <th>ID</th>
                    <th>ФИО</th>
                    <th>Участие</th>
                    <th>Гостей</th>
                    <th>Имена гостей</th>
                    <th>Второй день</th>
                    <th>Напитки</th>
                    <th>Дата</th>
                </tr>
            </thead>
            <tbody>
                {% for r in rows %}
                <tr>
                    <td>{{ r.id }}</td>
                    <td>{{ r.full_name }}</td>
                    <td>
                        {% if r.attendance == 'confirmed' %}
                            <span class="badge badge-confirmed">Приду</span>
                        {% elif r.attendance == 'declined' %}
                            <span class="badge badge-declined">Не приду</span>
                        {% else %}
                            <span class="badge badge-maybe">Возможно</span>
                        {% endif %}
                    </td>
                    <td>{{ r.guests_count }}</td>
                    <td>{{ r.guest_names }}</td>
                    <td>
                        {% if r.second_day == 'yes' %}
                            <span class="badge badge-confirmed">Приду</span>
                        {% elif r.second_day == 'no' %}
                            <span class="badge badge-declined">Не приду</span>
                        {% elif r.second_day == 'maybe' %}
                            <span class="badge badge-maybe">Возможно</span>
                        {% else %}
                            —
                        {% endif %}
                    </td>
                    <td>{{ r.drinks }}</td>
                    <td>{{ r.created_at }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% else %}
        <div class="empty">Пока нет ответов. Отправьте ссылку гостям! ☺️</div>
        {% endif %}
    </body>
    </html>
    """
    return render_template_string(
        html,
        rows=rows,
        total=total,
        confirmed=confirmed,
        declined=declined,
        maybe=maybe,
        total_guests=total_guests,
    )


# ───────────────────────────────────────────
# Запуск
# ───────────────────────────────────────────
if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "False").lower() in ("true", "1", "yes")
    app.run(host=host, port=port, debug=debug)
